"""Grade-report Excel downloads for teachers.

Issue #708 PR-2.

Two endpoints, both scoped to a classroom the requesting teacher owns:

- ``POST /classrooms/{classroom_id}/grade-report``
    Body: ``{"assignment_ids": [int]}``
    Returns an ``.xlsx`` summarising scores across the selected assignments.
    Assignments are grouped by ``score_category`` (聽力 / 閱讀 / 寫作 / 口說)
    matching the sample on the issue.

- ``POST /classrooms/{classroom_id}/student-grade-report``
    Body: ``{"student_ids": [int], "start_date": "YYYY-MM-DD"?, "end_date": "YYYY-MM-DD"?}``
    Returns an ``.xlsx`` with one sheet per selected student listing all of
    that student's assignments (optionally filtered by assignment
    ``created_at`` range) grouped by category, with an overall average.

The category for each assignment is read from ``assignment.score_category``
which is auto-resolved by ``utils.score_category.resolve_score_category``
(see issue #708 PR-1).
"""
from __future__ import annotations

import io
import logging
from datetime import date, datetime, timezone
from typing import List, Optional, Tuple
from urllib.parse import quote

from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import Response
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from pydantic import BaseModel, Field, field_validator
from sqlalchemy.orm import Session

from database import get_db
from models import (
    Assignment,
    Classroom,
    ClassroomStudent,
    Student,
    StudentAssignment,
    Teacher,
)
from routers.assignments.detail import _compute_interim_score
from .dependencies import get_current_teacher

logger = logging.getLogger(__name__)

router = APIRouter()


# Category display order on the class report (matches sample on the issue).
# Keys are score_category values stored in the DB.
_CATEGORY_ORDER: Tuple[str, ...] = ("listening", "reading", "writing", "speaking")
_CATEGORY_ZH = {
    "listening": "聽力",
    "reading": "閱讀",
    "writing": "寫作",
    "speaking": "口說",
}


# ---------------------------------------------------------------------------
# Request schemas
# ---------------------------------------------------------------------------


class ClassGradeReportRequest(BaseModel):
    assignment_ids: List[int] = Field(..., min_length=1)

    @field_validator("assignment_ids")
    @classmethod
    def _unique(cls, v: List[int]) -> List[int]:
        if len(set(v)) != len(v):
            raise ValueError("assignment_ids must not contain duplicates")
        return v


class StudentGradeReportRequest(BaseModel):
    student_ids: List[int] = Field(..., min_length=1)
    start_date: Optional[date] = None
    end_date: Optional[date] = None

    @field_validator("student_ids")
    @classmethod
    def _unique(cls, v: List[int]) -> List[int]:
        if len(set(v)) != len(v):
            raise ValueError("student_ids must not contain duplicates")
        return v


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _verify_classroom_ownership(
    classroom_id: int, teacher_id: int, db: Session
) -> Classroom:
    classroom = (
        db.query(Classroom)
        .filter(
            Classroom.id == classroom_id,
            Classroom.teacher_id == teacher_id,
            Classroom.is_active.is_(True),
        )
        .first()
    )
    if not classroom:
        raise HTTPException(
            status_code=404,
            detail="Classroom not found or you don't have permission",
        )
    return classroom


def _classroom_students(classroom_id: int, db: Session) -> List[Student]:
    return (
        db.query(Student)
        .join(ClassroomStudent, Student.id == ClassroomStudent.student_id)
        .filter(
            ClassroomStudent.classroom_id == classroom_id,
            ClassroomStudent.is_active.is_(True),
            Student.is_active.is_(True),
        )
        .order_by(Student.student_number.asc().nullslast(), Student.id.asc())
        .all()
    )


def _category_key(assignment: Assignment) -> str:
    """Return the category bucket key for an assignment.

    Falls back to ``"writing"`` when the assignment has no category. After
    PR-1 (#744, #746) every row has a value, but this keeps the report
    robust against any future row that slipped through.
    """
    key = (assignment.score_category or "").strip().lower()
    return key if key in _CATEGORY_ZH else "writing"


def _final_score(sa: Optional[StudentAssignment], assignment: Assignment, db: Session):
    """Return a numeric score (or ``None``) for one (student, assignment) cell.

    Sample reports show integer scores, so we round to the nearest integer.
    """
    if sa is None:
        return None
    if sa.score is not None:
        return int(round(float(sa.score)))
    interim = _compute_interim_score(sa, assignment, db)
    return int(round(float(interim))) if interim is not None else None


def _avg(scores: List[float]) -> Optional[float]:
    return round(sum(scores) / len(scores), 1) if scores else None


def _today_yyyymmdd() -> str:
    return datetime.now(timezone.utc).strftime("%Y%m%d")


def _xlsx_response(wb: Workbook, filename: str) -> Response:
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    # RFC 5987: filename* lets us send a UTF-8 filename with full fidelity,
    # while filename= provides a safe ASCII fallback for older browsers.
    ascii_fallback = filename.encode("ascii", errors="ignore").decode() or "report.xlsx"
    disposition = (
        f'attachment; filename="{ascii_fallback}"; '
        f"filename*=UTF-8''{quote(filename)}"
    )
    return Response(
        content=buf.getvalue(),
        media_type=(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ),
        headers={"Content-Disposition": disposition},
    )


def _sanitize_sheet_title(raw: str) -> str:
    """openpyxl rejects ``[]:*?/\\`` and titles longer than 31 chars."""
    bad = set("[]:*?/\\")
    cleaned = "".join(ch for ch in raw if ch not in bad).strip() or "Sheet"
    return cleaned[:31]


# ---------------------------------------------------------------------------
# Class grade report (班級成績單)
# ---------------------------------------------------------------------------


def _build_class_workbook(
    classroom: Classroom,
    assignments: List[Assignment],
    students: List[Student],
    db: Session,
) -> Workbook:
    """Build the class grade report workbook.

    Layout matches the sample on issue #708:

    ::

        Row 1: <ClassName> 班成績總覽           (merged across all columns)
        Row 2: 下載日期：YYYY-MM-DD             (merged across all columns)
        Row 3: 座號|姓名|個人平均|<category banner cells>|班級平均
        Row 4: (blank x3)|<assignment title cells>|(blank)
        Row 5: (blank x3)|<assignment date cells>|(blank)
        Row 6+: per-student score rows
        Last:  class-average row
    """
    # Group assignments by category, preserving stable order within a group.
    groups: dict[str, List[Assignment]] = {k: [] for k in _CATEGORY_ORDER}
    for a in assignments:
        groups[_category_key(a)].append(a)

    # Drop empty groups so we don't leave dead columns.
    non_empty = [
        (cat, items) for cat in _CATEGORY_ORDER for items in (groups[cat],) if items
    ]

    # Flat ordered list of assignments, paired with their category, plus the
    # starting column index for each category group (used for merging).
    ordered: List[Tuple[str, Assignment]] = []
    cat_spans: List[Tuple[str, int, int]] = []  # (category, start_col, end_col)
    col = 4  # 座號(1), 姓名(2), 個人平均(3), then assignments from col 4
    for cat, items in non_empty:
        start = col
        for a in items:
            ordered.append((cat, a))
            col += 1
        cat_spans.append((cat, start, col - 1))

    wb = Workbook()
    ws = wb.active
    ws.title = _sanitize_sheet_title(f"{classroom.name}成績總覽")

    total_cols = col - 1  # last assignment column

    # --- Row 1: title ---
    ws.cell(row=1, column=1, value=f"{classroom.name} 班成績總覽")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
    ws.cell(row=1, column=1).font = Font(bold=True, size=14)
    ws.cell(row=1, column=1).alignment = Alignment(horizontal="center")

    # --- Row 2: download date ---
    today_iso = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    ws.cell(row=2, column=1, value=f"下載日期：{today_iso}")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=total_cols)
    ws.cell(row=2, column=1).alignment = Alignment(horizontal="left")

    # --- Row 3: top header (with category banners merged) ---
    headers_row3 = ["座號", "姓名", "個人平均"]
    ws.cell(row=3, column=1, value=headers_row3[0])
    ws.cell(row=3, column=2, value=headers_row3[1])
    ws.cell(row=3, column=3, value=headers_row3[2])
    for cat, start, end in cat_spans:
        ws.cell(row=3, column=start, value=_CATEGORY_ZH[cat])
        if start != end:
            ws.merge_cells(start_row=3, start_column=start, end_row=3, end_column=end)
        ws.cell(row=3, column=start).alignment = Alignment(horizontal="center")
    # Vertically merge the static columns (座號/姓名/個人平均) over
    # rows 3..5 so they read as a single header cell.
    for static_col in (1, 2, 3):
        ws.merge_cells(
            start_row=3,
            start_column=static_col,
            end_row=5,
            end_column=static_col,
        )
        ws.cell(row=3, column=static_col).alignment = Alignment(
            horizontal="center", vertical="center"
        )

    # --- Row 4: assignment titles ---
    for idx, (_, a) in enumerate(ordered):
        ws.cell(row=4, column=4 + idx, value=a.title)
        ws.cell(row=4, column=4 + idx).alignment = Alignment(
            horizontal="center", wrap_text=True
        )

    # --- Row 5: assignment dates (created_at, treated as 派發日期) ---
    for idx, (_, a) in enumerate(ordered):
        d = a.created_at.date().isoformat() if a.created_at else ""
        ws.cell(row=5, column=4 + idx, value=d)
        ws.cell(row=5, column=4 + idx).alignment = Alignment(horizontal="center")

    # --- Pre-load student assignments to avoid N+1 ---
    assignment_ids = [a.id for a in assignments]
    sa_rows = (
        db.query(StudentAssignment)
        .filter(
            StudentAssignment.assignment_id.in_(assignment_ids),
            StudentAssignment.is_active.is_(True),
        )
        .all()
    )
    sa_by_pair = {(sa.student_id, sa.assignment_id): sa for sa in sa_rows}

    # --- Data rows ---
    all_scores: List[float] = []
    per_assignment_scores: dict[int, List[float]] = {a.id: [] for a in assignments}

    first_data_row = 6
    for student_idx, student in enumerate(students):
        row = first_data_row + student_idx
        ws.cell(row=row, column=1, value=student.student_number or "")
        ws.cell(row=row, column=2, value=student.name or "")
        student_scores: List[float] = []

        for col_idx, (_, a) in enumerate(ordered):
            sa = sa_by_pair.get((student.id, a.id))
            score = _final_score(sa, a, db)
            if score is not None:
                student_scores.append(float(score))
                per_assignment_scores[a.id].append(float(score))
                all_scores.append(float(score))
                ws.cell(row=row, column=4 + col_idx, value=score)
            # else: leave blank (per user spec)
            ws.cell(row=row, column=4 + col_idx).alignment = Alignment(
                horizontal="center"
            )

        personal_avg = _avg(student_scores)
        if personal_avg is not None:
            ws.cell(row=row, column=3, value=personal_avg)
        ws.cell(row=row, column=3).alignment = Alignment(horizontal="center")

    # --- Footer row: 班級平均 ---
    footer_row = first_data_row + len(students)
    ws.cell(row=footer_row, column=1, value="班級平均")
    ws.merge_cells(
        start_row=footer_row, start_column=1, end_row=footer_row, end_column=2
    )
    ws.cell(row=footer_row, column=1).font = Font(bold=True)
    ws.cell(row=footer_row, column=1).alignment = Alignment(horizontal="center")

    overall_avg = _avg(all_scores)
    if overall_avg is not None:
        ws.cell(row=footer_row, column=3, value=overall_avg)
    for col_idx, (_, a) in enumerate(ordered):
        per_a = _avg(per_assignment_scores[a.id])
        if per_a is not None:
            ws.cell(row=footer_row, column=4 + col_idx, value=per_a)
        ws.cell(row=footer_row, column=4 + col_idx).alignment = Alignment(
            horizontal="center"
        )
    ws.cell(row=footer_row, column=3).alignment = Alignment(horizontal="center")

    # --- Light styling: header background + borders ---
    header_fill = PatternFill("solid", fgColor="F2F2F2")
    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for header_row in (3, 4, 5):
        for c in range(1, total_cols + 1):
            cell = ws.cell(row=header_row, column=c)
            cell.fill = header_fill
            cell.font = cell.font.copy(bold=True)
            cell.border = border
    for r in range(first_data_row, footer_row + 1):
        for c in range(1, total_cols + 1):
            ws.cell(row=r, column=c).border = border

    # --- Column widths ---
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 10
    for c in range(4, total_cols + 1):
        ws.column_dimensions[get_column_letter(c)].width = 16
    ws.row_dimensions[1].height = 24
    ws.row_dimensions[4].height = 30

    return wb


@router.post("/classrooms/{classroom_id}/grade-report")
async def class_grade_report(
    classroom_id: int,
    request: ClassGradeReportRequest,
    db: Session = Depends(get_db),
    current_teacher: Teacher = Depends(get_current_teacher),
):
    classroom = _verify_classroom_ownership(classroom_id, current_teacher.id, db)

    assignments = (
        db.query(Assignment)
        .filter(
            Assignment.id.in_(request.assignment_ids),
            Assignment.classroom_id == classroom_id,
            Assignment.teacher_id == current_teacher.id,
            Assignment.is_active.is_(True),
        )
        .order_by(Assignment.created_at.asc(), Assignment.id.asc())
        .all()
    )
    if len(assignments) != len(request.assignment_ids):
        raise HTTPException(
            status_code=404,
            detail=(
                "Some assignments were not found in this classroom or are "
                "inaccessible"
            ),
        )

    students = _classroom_students(classroom_id, db)
    wb = _build_class_workbook(classroom, assignments, students, db)
    filename = f"{classroom.name}_成績總覽_{_today_yyyymmdd()}.xlsx"
    return _xlsx_response(wb, filename)


# ---------------------------------------------------------------------------
# Student grade report (學生成績單)
# ---------------------------------------------------------------------------


def _build_student_workbook(
    classroom: Classroom,
    students: List[Student],
    start_date: Optional[date],
    end_date: Optional[date],
    db: Session,
) -> Workbook:
    """One sheet per student. Sheet layout (matches sample):

    ::

        Row 1: 學生成績單　<class> 班　<num> 號　<name>
        Row 2: 下載日期：YYYY-MM-DD
        Row 3: 類別 | 作業名稱 | 派發日期 | 分數
        Then for each non-empty category:
            Row:  ▌ <category>
            Rows: <category> | <title> | <date> | <score>
        Last: 總平均 | | | <avg>
    """
    # Pull all assignments owned by this teacher for this classroom in one go,
    # so each student sheet just filters in memory.
    teacher_id = classroom.teacher_id
    q = db.query(Assignment).filter(
        Assignment.classroom_id == classroom.id,
        Assignment.teacher_id == teacher_id,
        Assignment.is_active.is_(True),
    )
    if start_date is not None:
        q = q.filter(
            Assignment.created_at >= datetime.combine(start_date, datetime.min.time())
        )
    if end_date is not None:
        # Inclusive end_date — bump to the start of next day.
        q = q.filter(
            Assignment.created_at < datetime.combine(end_date, datetime.max.time())
        )
    classroom_assignments = q.order_by(Assignment.created_at.asc()).all()

    sa_rows = (
        db.query(StudentAssignment)
        .filter(
            StudentAssignment.assignment_id.in_([a.id for a in classroom_assignments])
            if classroom_assignments
            else False,
            StudentAssignment.is_active.is_(True),
        )
        .all()
    )
    sa_by_pair = {(sa.student_id, sa.assignment_id): sa for sa in sa_rows}

    wb = Workbook()
    # Remove the default blank sheet; we'll add one per student.
    wb.remove(wb.active)

    today_iso = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill("solid", fgColor="F2F2F2")
    banner_fill = PatternFill("solid", fgColor="E8EEF7")

    for student in students:
        title = _sanitize_sheet_title(
            f"{classroom.name}_{student.student_number or ''}_{student.name or ''}"
        )
        ws = wb.create_sheet(title=title)

        # Row 1: header
        ws.cell(
            row=1,
            column=1,
            value=(
                f"學生成績單　{classroom.name} 班　"
                f"{student.student_number or ''} 號　{student.name or ''}"
            ),
        )
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)
        ws.cell(row=1, column=1).font = Font(bold=True, size=13)
        ws.cell(row=1, column=1).alignment = Alignment(horizontal="left")

        # Row 2: download date
        ws.cell(row=2, column=1, value=f"下載日期：{today_iso}")
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=4)

        # Row 3: header
        for c, h in enumerate(("類別", "作業名稱", "派發日期", "分數"), start=1):
            cell = ws.cell(row=3, column=c, value=h)
            cell.fill = header_fill
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")
            cell.border = border

        # Bucket assignments by category and write rows.
        groups: dict[str, List[Assignment]] = {k: [] for k in _CATEGORY_ORDER}
        for a in classroom_assignments:
            groups[_category_key(a)].append(a)

        cur_row = 4
        scores: List[float] = []
        for cat in _CATEGORY_ORDER:
            items = groups[cat]
            if not items:
                continue
            # Banner row
            banner = ws.cell(row=cur_row, column=1, value=f"▌ {_CATEGORY_ZH[cat]}")
            ws.merge_cells(
                start_row=cur_row, start_column=1, end_row=cur_row, end_column=4
            )
            banner.font = Font(bold=True)
            banner.fill = banner_fill
            banner.alignment = Alignment(horizontal="left")
            for c in range(1, 5):
                ws.cell(row=cur_row, column=c).border = border
            cur_row += 1

            for a in items:
                sa = sa_by_pair.get((student.id, a.id))
                score = _final_score(sa, a, db)
                d = a.created_at.date().isoformat() if a.created_at else ""
                ws.cell(row=cur_row, column=1, value=_CATEGORY_ZH[cat])
                ws.cell(row=cur_row, column=2, value=a.title)
                ws.cell(row=cur_row, column=3, value=d)
                if score is not None:
                    ws.cell(row=cur_row, column=4, value=score)
                    scores.append(float(score))
                # else: blank cell
                for c in range(1, 5):
                    cell = ws.cell(row=cur_row, column=c)
                    cell.border = border
                    if c >= 3:
                        cell.alignment = Alignment(horizontal="center")
                cur_row += 1

        # Footer: 總平均
        avg = _avg(scores)
        ws.cell(row=cur_row, column=1, value="總平均")
        ws.merge_cells(start_row=cur_row, start_column=1, end_row=cur_row, end_column=3)
        ws.cell(row=cur_row, column=1).font = Font(bold=True)
        ws.cell(row=cur_row, column=1).alignment = Alignment(horizontal="center")
        if avg is not None:
            ws.cell(row=cur_row, column=4, value=avg)
        ws.cell(row=cur_row, column=4).alignment = Alignment(horizontal="center")
        for c in range(1, 5):
            ws.cell(row=cur_row, column=c).border = border

        # Column widths
        ws.column_dimensions["A"].width = 14
        ws.column_dimensions["B"].width = 28
        ws.column_dimensions["C"].width = 14
        ws.column_dimensions["D"].width = 10

    return wb


@router.post("/classrooms/{classroom_id}/student-grade-report")
async def student_grade_report(
    classroom_id: int,
    request: StudentGradeReportRequest,
    db: Session = Depends(get_db),
    current_teacher: Teacher = Depends(get_current_teacher),
):
    if (
        request.start_date is not None
        and request.end_date is not None
        and request.start_date > request.end_date
    ):
        raise HTTPException(
            status_code=422, detail="start_date must be on or before end_date"
        )

    classroom = _verify_classroom_ownership(classroom_id, current_teacher.id, db)

    # Validate all student IDs belong to this classroom.
    students = (
        db.query(Student)
        .join(ClassroomStudent, Student.id == ClassroomStudent.student_id)
        .filter(
            ClassroomStudent.classroom_id == classroom_id,
            ClassroomStudent.is_active.is_(True),
            Student.is_active.is_(True),
            Student.id.in_(request.student_ids),
        )
        .order_by(Student.student_number.asc().nullslast(), Student.id.asc())
        .all()
    )
    if len(students) != len(request.student_ids):
        raise HTTPException(
            status_code=404,
            detail="Some students were not found in this classroom",
        )

    wb = _build_student_workbook(
        classroom, students, request.start_date, request.end_date, db
    )

    if len(students) == 1:
        s = students[0]
        filename = (
            f"{classroom.name}_{s.student_number or ''}_{s.name or ''}_"
            f"{_today_yyyymmdd()}.xlsx"
        )
    else:
        filename = f"{classroom.name}_學生成績單_{_today_yyyymmdd()}.xlsx"
    return _xlsx_response(wb, filename)


# Re-export so the package can register this router.
__all__ = ["router"]
