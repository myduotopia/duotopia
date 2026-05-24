"""Integration tests for the grade-report endpoints (issue #708)."""
from __future__ import annotations

import io
import zipfile
from datetime import datetime, timedelta, timezone

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlalchemy import create_engine, event
from sqlalchemy.orm import sessionmaker

from auth import get_password_hash
from database import Base, get_db
from main import app
from models import (
    Assignment,
    Classroom,
    ClassroomStudent,
    Student,
    StudentAssignment,
    SubscriptionPeriod,
    Teacher,
)


SQLALCHEMY_DATABASE_URL = "sqlite:///./test_grade_reports.db"
engine = create_engine(
    SQLALCHEMY_DATABASE_URL,
    connect_args={"check_same_thread": False},
    pool_pre_ping=True,
)


@event.listens_for(engine, "connect")
def _enable_fk(dbapi_conn, _record):
    cur = dbapi_conn.cursor()
    cur.execute("PRAGMA foreign_keys=ON")
    cur.close()


TestingSessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=engine)


def _override_get_db():
    db = TestingSessionLocal()
    try:
        yield db
    finally:
        db.close()


app.dependency_overrides[get_db] = _override_get_db
client = TestClient(app)


@pytest.fixture(scope="function")
def fresh_db():
    Base.metadata.create_all(bind=engine)
    yield
    Base.metadata.drop_all(bind=engine)


def _login(email: str = "t1@test.com", password: str = "pw") -> dict:
    res = client.post(
        "/api/auth/teacher/login",
        json={"email": email, "password": password},
    )
    assert res.status_code == 200, res.text
    return {"Authorization": f"Bearer {res.json()['access_token']}"}


@pytest.fixture
def seeded(fresh_db):
    """Seed: teacher T1 with classroom C1 containing two students and three
    assignments across categories. T2 owns a separate classroom for the
    permission test."""
    db = TestingSessionLocal()
    now = datetime.now(timezone.utc)

    # Two teachers, each with a subscription period so they can call the API.
    for tid, email in [(1, "t1@test.com"), (2, "t2@test.com")]:
        db.add(
            Teacher(
                id=tid,
                name=f"Teacher {tid}",
                email=email,
                password_hash=get_password_hash("pw"),
                email_verified=True,
                is_active=True,
            )
        )
    db.commit()
    for tid in (1, 2):
        db.add(
            SubscriptionPeriod(
                teacher_id=tid,
                plan_name="Tutor Teachers",
                amount_paid=299,
                quota_total=2000,
                quota_used=0,
                start_date=now,
                end_date=now + timedelta(days=30),
                payment_method="credit_card",
            )
        )
    db.commit()

    # Classrooms
    db.add(Classroom(id=1, name="301", teacher_id=1, is_active=True))
    db.add(Classroom(id=2, name="OtherClass", teacher_id=2, is_active=True))
    db.commit()

    # Students in C1 (T1's classroom): 01 王小明, 02 林小花
    db.add(
        Student(
            id=11,
            name="王小明",
            student_number="01",
            email="s11@test.com",
            password_hash=get_password_hash("pw"),
            is_active=True,
        )
    )
    db.add(
        Student(
            id=12,
            name="林小花",
            student_number="02",
            email="s12@test.com",
            password_hash=get_password_hash("pw"),
            is_active=True,
        )
    )
    db.commit()
    db.add(ClassroomStudent(classroom_id=1, student_id=11, is_active=True))
    db.add(ClassroomStudent(classroom_id=1, student_id=12, is_active=True))
    db.commit()

    # Three assignments owned by T1 in C1 across categories.
    # Dates: 30 days ago, 20 days ago, 10 days ago — used in date-range test.
    base = now - timedelta(days=30)
    a_listening = Assignment(
        id=101,
        title="Unit 1 Listening Quiz",
        classroom_id=1,
        teacher_id=1,
        practice_mode="rearrangement",
        play_audio=True,
        score_category="listening",
        is_active=True,
        created_at=base,
    )
    a_reading = Assignment(
        id=102,
        title="Chapter 1 Reading",
        classroom_id=1,
        teacher_id=1,
        practice_mode="word_cloze",
        play_audio=False,
        score_category="reading",
        is_active=True,
        created_at=base + timedelta(days=10),
    )
    a_speaking = Assignment(
        id=103,
        title="Self-Introduction",
        classroom_id=1,
        teacher_id=1,
        practice_mode="word_reading",
        play_audio=False,
        score_category="speaking",
        is_active=True,
        created_at=base + timedelta(days=20),
    )
    for a in (a_listening, a_reading, a_speaking):
        db.add(a)
    db.commit()

    # StudentAssignment scores. 王小明 has all three; 林小花 only has two.
    db.add(
        StudentAssignment(
            assignment_id=101,
            student_id=11,
            teacher_id=1,
            classroom_id=1,
            title="Unit 1 Listening Quiz",
            score=88,
            is_active=True,
        )
    )
    db.add(
        StudentAssignment(
            assignment_id=102,
            student_id=11,
            teacher_id=1,
            classroom_id=1,
            title="Chapter 1 Reading",
            score=76,
            is_active=True,
        )
    )
    db.add(
        StudentAssignment(
            assignment_id=103,
            student_id=11,
            teacher_id=1,
            classroom_id=1,
            title="Self-Introduction",
            score=95,
            is_active=True,
        )
    )
    db.add(
        StudentAssignment(
            assignment_id=101,
            student_id=12,
            teacher_id=1,
            classroom_id=1,
            title="Unit 1 Listening Quiz",
            score=75,
            is_active=True,
        )
    )
    # 林小花 has no row for assignment 102 (intentional — leave blank).
    db.add(
        StudentAssignment(
            assignment_id=103,
            student_id=12,
            teacher_id=1,
            classroom_id=1,
            title="Self-Introduction",
            score=88,
            is_active=True,
        )
    )
    db.commit()
    db.close()
    return {
        "classroom_id": 1,
        "other_classroom_id": 2,
        "assignment_ids": [101, 102, 103],
        "student_ids": [11, 12],
        "assignment_listening_id": 101,
        "assignment_reading_id": 102,
        "assignment_speaking_id": 103,
    }


def _load(content: bytes):
    return load_workbook(io.BytesIO(content), data_only=True)


# ---------------------------------------------------------------------------
# Class grade report
# ---------------------------------------------------------------------------


def test_class_grade_report_happy_path(seeded):
    headers = _login()
    res = client.post(
        f"/api/teachers/classrooms/{seeded['classroom_id']}/grade-report",
        headers=headers,
        json={"assignment_ids": seeded["assignment_ids"]},
    )
    assert res.status_code == 200, res.text
    assert "spreadsheetml" in res.headers["content-type"]
    assert "Content-Disposition" in res.headers
    assert "filename*=UTF-8''" in res.headers["Content-Disposition"]

    wb = _load(res.content)
    ws = wb.active
    # Title row mentions class name.
    assert "301" in str(ws.cell(row=1, column=1).value)
    # Static headers
    assert ws.cell(row=3, column=1).value == "座號"
    assert ws.cell(row=3, column=2).value == "姓名"
    assert ws.cell(row=3, column=3).value == "個人平均"
    # Categories appear in 聽力 → 閱讀 → 寫作 → 口說 order
    banner_cells = [
        str(c.value) for c in ws[3] if c.value and c.value not in ("座號", "姓名", "個人平均")
    ]
    assert banner_cells == ["聽力", "閱讀", "口說"]
    # The rightmost 班級平均 column was removed (issue #708 follow-up):
    # the rightmost header cell must be the last assignment's category banner
    # or an assignment cell — never "班級平均".
    row3_values = [c.value for c in ws[3] if c.value]
    assert "班級平均" not in row3_values
    # 王小明 row: student_number 01, has scores 88 / 76 / 95
    student_rows = [r for r in ws.iter_rows(min_row=6, max_row=7, values_only=True)]
    names = [r[1] for r in student_rows]
    assert "王小明" in names
    assert "林小花" in names
    # Footer: 班級平均 label still anchors the bottom row in cols 1-2,
    # and the rightmost footer cell is now a per-assignment average (or
    # blank), never the duplicated overall average.
    last_row = list(ws.iter_rows(values_only=True))[-1]
    assert last_row[0] == "班級平均"


def test_class_grade_report_rejects_assignments_in_other_classroom(seeded):
    """T1 cannot include an ID that belongs to T2's classroom."""
    # Seed an extra assignment owned by T2 in their classroom.
    db = TestingSessionLocal()
    db.add(
        Assignment(
            id=999,
            title="Other classroom assignment",
            classroom_id=seeded["other_classroom_id"],
            teacher_id=2,
            practice_mode="reading",
            play_audio=False,
            score_category="speaking",
            is_active=True,
        )
    )
    db.commit()
    db.close()

    headers = _login()
    res = client.post(
        f"/api/teachers/classrooms/{seeded['classroom_id']}/grade-report",
        headers=headers,
        json={"assignment_ids": [seeded["assignment_listening_id"], 999]},
    )
    assert res.status_code == 404


def test_class_grade_report_404_for_classroom_not_owned(seeded):
    headers = _login()  # T1
    res = client.post(
        f"/api/teachers/classrooms/{seeded['other_classroom_id']}/grade-report",
        headers=headers,
        json={"assignment_ids": [seeded["assignment_listening_id"]]},
    )
    assert res.status_code == 404


def test_class_grade_report_rejects_empty_selection(seeded):
    headers = _login()
    res = client.post(
        f"/api/teachers/classrooms/{seeded['classroom_id']}/grade-report",
        headers=headers,
        json={"assignment_ids": []},
    )
    assert res.status_code == 422


def test_class_grade_report_unauthenticated_returns_401(seeded):
    res = client.post(
        f"/api/teachers/classrooms/{seeded['classroom_id']}/grade-report",
        json={"assignment_ids": seeded["assignment_ids"]},
    )
    assert res.status_code == 401


# ---------------------------------------------------------------------------
# Student grade report (ZIP-of-xlsx contract, issue #708 follow-up)
# ---------------------------------------------------------------------------


def _open_zip(content: bytes) -> zipfile.ZipFile:
    return zipfile.ZipFile(io.BytesIO(content))


def test_student_grade_report_happy_path_returns_zip_per_enrolled_student(seeded):
    """Selecting all three assignments produces a zip with one xlsx per
    enrolled student (2 in this fixture), each containing only the selected
    assignments and the right 總平均."""
    headers = _login()
    res = client.post(
        f"/api/teachers/classrooms/{seeded['classroom_id']}/student-grade-report",
        headers=headers,
        json={"assignment_ids": seeded["assignment_ids"]},
    )
    assert res.status_code == 200, res.text
    assert res.headers["content-type"] == "application/zip"
    assert "Content-Disposition" in res.headers
    assert "filename*=UTF-8''" in res.headers["Content-Disposition"]
    # Zip filename pattern: <class>_學生成績單_<YYYYMMDD>.zip
    assert "301_%E5%AD%B8%E7%94%9F" in res.headers["Content-Disposition"]
    assert res.headers["Content-Disposition"].endswith(".zip")

    zf = _open_zip(res.content)
    names = zf.namelist()
    assert len(names) == 2  # 王小明 + 林小花
    assert any("王小明" in n for n in names)
    assert any("林小花" in n for n in names)
    # Inner files end in .xlsx
    assert all(n.endswith(".xlsx") for n in names)

    # 王小明: 88 / 76 / 95 → 總平均 86.3
    wang_name = next(n for n in names if "王小明" in n)
    wb = load_workbook(io.BytesIO(zf.read(wang_name)), data_only=True)
    ws = wb.active
    headers_row = [ws.cell(row=3, column=c).value for c in range(1, 5)]
    assert headers_row == ["類別", "作業名稱", "派發日期", "分數"]
    rows = list(ws.iter_rows(values_only=True))
    assert rows[-1][0] == "總平均"
    assert rows[-1][3] == pytest.approx(86.3)


def test_student_grade_report_subset_of_assignments_changes_average(seeded):
    """Selecting only the reading assignment (王小明 scored 76) should make
    that student's 總平均 == 76 — proves the per-student xlsx really is scoped
    to the selection rather than always covering every assignment."""
    headers = _login()
    res = client.post(
        f"/api/teachers/classrooms/{seeded['classroom_id']}/student-grade-report",
        headers=headers,
        json={"assignment_ids": [seeded["assignment_reading_id"]]},
    )
    assert res.status_code == 200, res.text

    zf = _open_zip(res.content)
    wang_name = next(n for n in zf.namelist() if "王小明" in n)
    ws = load_workbook(io.BytesIO(zf.read(wang_name)), data_only=True).active
    all_text = " ".join(str(c.value) for row in ws.iter_rows() for c in row if c.value)
    assert "▌ 閱讀" in all_text
    assert "▌ 聽力" not in all_text
    assert "▌ 口說" not in all_text
    rows = list(ws.iter_rows(values_only=True))
    assert rows[-1][0] == "總平均"
    assert rows[-1][3] == pytest.approx(76)


def test_student_grade_report_one_student_classroom_still_returns_zip(fresh_db):
    """A classroom with exactly one student should still come back as a ZIP
    containing one file — frontend treats every response as a ZIP."""
    db = TestingSessionLocal()
    now = datetime.now(timezone.utc)
    db.add(
        Teacher(
            id=1,
            name="Solo Teacher",
            email="solo@test.com",
            password_hash=get_password_hash("pw"),
            email_verified=True,
            is_active=True,
        )
    )
    db.commit()
    db.add(
        SubscriptionPeriod(
            teacher_id=1,
            plan_name="Tutor Teachers",
            amount_paid=299,
            quota_total=2000,
            quota_used=0,
            start_date=now,
            end_date=now + timedelta(days=30),
            payment_method="credit_card",
        )
    )
    db.add(Classroom(id=10, name="SoloClass", teacher_id=1, is_active=True))
    db.commit()
    db.add(
        Student(
            id=20,
            name="OnlyKid",
            student_number="01",
            email="only@test.com",
            password_hash=get_password_hash("pw"),
            is_active=True,
        )
    )
    db.commit()
    db.add(ClassroomStudent(classroom_id=10, student_id=20, is_active=True))
    db.add(
        Assignment(
            id=200,
            title="Only Quiz",
            classroom_id=10,
            teacher_id=1,
            practice_mode="word_reading",
            play_audio=False,
            score_category="speaking",
            is_active=True,
            created_at=now,
        )
    )
    db.commit()
    db.close()

    headers = _login(email="solo@test.com")
    res = client.post(
        "/api/teachers/classrooms/10/student-grade-report",
        headers=headers,
        json={"assignment_ids": [200]},
    )
    assert res.status_code == 200, res.text
    assert res.headers["content-type"] == "application/zip"
    zf = _open_zip(res.content)
    assert len(zf.namelist()) == 1
    assert "OnlyKid" in zf.namelist()[0]


def test_student_grade_report_rejects_assignments_in_other_classroom(seeded):
    """T1 cannot include an assignment id that belongs to T2's classroom."""
    db = TestingSessionLocal()
    db.add(
        Assignment(
            id=999,
            title="Other classroom assignment",
            classroom_id=seeded["other_classroom_id"],
            teacher_id=2,
            practice_mode="reading",
            play_audio=False,
            score_category="speaking",
            is_active=True,
        )
    )
    db.commit()
    db.close()

    headers = _login()
    res = client.post(
        f"/api/teachers/classrooms/{seeded['classroom_id']}/student-grade-report",
        headers=headers,
        json={"assignment_ids": [seeded["assignment_listening_id"], 999]},
    )
    assert res.status_code == 404


def test_student_grade_report_404_for_classroom_not_owned(seeded):
    headers = _login()  # T1
    res = client.post(
        f"/api/teachers/classrooms/{seeded['other_classroom_id']}/student-grade-report",
        headers=headers,
        json={"assignment_ids": [seeded["assignment_listening_id"]]},
    )
    assert res.status_code == 404


def test_student_grade_report_rejects_empty_selection(seeded):
    headers = _login()
    res = client.post(
        f"/api/teachers/classrooms/{seeded['classroom_id']}/student-grade-report",
        headers=headers,
        json={"assignment_ids": []},
    )
    assert res.status_code == 422


def test_student_grade_report_unauthenticated_returns_401(seeded):
    res = client.post(
        f"/api/teachers/classrooms/{seeded['classroom_id']}/student-grade-report",
        json={"assignment_ids": seeded["assignment_ids"]},
    )
    assert res.status_code == 401
